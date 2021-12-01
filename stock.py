#!/usr/bin/env python
#coding=utf-8

import easyquotation
import openpyxl

class Stock(object):
    def __init__(self, source="sina"):
        self.quotation = easyquotation.use(source)
        self.quotation.market_snapshot(prefix=True)
    
    def get_stock_price(self, stock_code):
        if isinstance(stock_code, str) != True:
            raise Exception("股票代码类型错误")
        return self.quotation.real(stock_code)[stock_code]["now"]

class Sheet(object):
    def __init__(self, name="账户明细.xlsx"):
        self.name = name
        self.book = openpyxl.load_workbook(name)
        # 获取第一张表
        self.sheet = self.book[self.book.sheetnames[0]]
        self.rows = self.sheet.max_row
        self.cols = self.sheet.max_column

    def get_value(self, postion):
        if isinstance(postion, str) != True:
            raise Exception("表格位置参数类型错误")
        return self.sheet[postion].value

    def set_value(self, postion, value):
        if isinstance(postion, str) != True:
            raise Exception("表格位置参数类型错误")
        self.sheet[postion].value = value

    def save(self):
        self.book.save(self.name)

    def close(self):
        self.book.save(self.name)
        self.book.close()

if __name__ == "__main__":
    source = input("请输入股票源地址（默认使用\"sina\"）：")
    name = input("请输入需要更新的表格（默认更新\"账户明细.xlsx\"）：")
    if source == "":
        source = "sina"
    if name == "":
        name = "账户明细.xlsx"
    try:
        # 实例化Stock类
        stock = Stock(source)
        # 实例化excel表格
        sheet = Sheet()
        # 更新账户持仓股票的股价
        for i in range(sheet.rows):
            # 剔除第一行和最后一行
            if(i == 0 or i == (sheet.rows - 1) or i == (sheet.rows - 2)):
                continue
            # 获取股票代码
            stock_code = sheet.get_value("B{0}".format(i + 1))
            # 获取最新股价
            stock_price = stock.get_stock_price("{0}".format(stock_code))
            # 更新表格的股价
            sheet.set_value("C{0}".format(i + 1), stock_price)
    finally:
        if sheet != None:
            # 关闭表格
            sheet.close()

